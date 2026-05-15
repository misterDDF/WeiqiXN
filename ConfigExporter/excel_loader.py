# -*- coding: utf-8 -*-
"""
Excel 读取与缓存层
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List

try:
    import openpyxl
except ImportError:
    raise ImportError("缺少必要的依赖库 openpyxl，请先运行 setup.bat")

from excel_common import SheetCache


class ExcelLoader:
    def __init__(self, excel_path: str | Path):
        self.excel_path = Path(excel_path)
        self.wb = None
        self.sheet_cache: Dict[str, SheetCache] = {}

    def load(self) -> None:
        if not self.excel_path.exists():
            raise FileNotFoundError(f"文件不存在: {self.excel_path}")

        if self.wb is not None:
            self.close()

        self.wb = openpyxl.load_workbook(self.excel_path)
        self.sheet_cache.clear()

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            self.sheet_cache[sheet_name] = SheetCache(
                sheet_name=sheet_name,
                rows=rows,
                max_col=ws.max_column,
                max_row=ws.max_row,
            )

    def close(self) -> None:
        if self.wb:
            self.wb.close()
            self.wb = None
        self.sheet_cache.clear()

    def get_valid_sheets(self) -> List[str]:
        if self.wb is None:
            raise RuntimeError("工作簿未加载")
        return [s for s in self.wb.sheetnames if not s.startswith('#')]

    def get_sheet_cache(self, sheet_name: str) -> SheetCache:
        if sheet_name not in self.sheet_cache:
            raise ValueError(f"未找到工作表 '{sheet_name}'")
        return self.sheet_cache[sheet_name]
